import openpyxl
import csv
import os
import xlrd

def _process_xls(file_path: str) -> str:
    """
    Processes a legacy .xls file using xlrd and extracts its content as text.
    """
    try:
        workbook = xlrd.open_workbook(file_path)
        text_content = []
        for sheet in workbook.sheets():
            text_content.append(f"--- Sheet: {sheet.name} ---")
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                text_content.append("\t".join(row_values))
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error reading XLS {file_path}: {e}")
        return ""

def _process_xlsx(file_path: str) -> str:
    """
    Processes an .xlsx file and extracts its content as text.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        text_content = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                text_content.append("\t".join(row_values))
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error reading XLSX {file_path}: {e}")
        return ""

def _process_csv(file_path: str) -> str:
    """
    Processes a .csv file and converts it into a Markdown table.
    """
    try:
        with open(file_path, mode='r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            
            # Read header and data
            header = next(reader)
            data = list(reader)
            
            # Create Markdown table
            markdown_table = []
            markdown_table.append(f"| {' | '.join(header)} |")
            markdown_table.append(f"|{':---:|' * len(header)}")
            
            for row in data:
                markdown_table.append(f"| {' | '.join(row)} |")
                
            return "\n".join(markdown_table)
    except Exception as e:
        print(f"Error reading CSV {file_path}: {e}")
        return ""

def process(file_path: str) -> str:
    """
    Processes a spreadsheet file (.xlsx, .xls, or .csv) based on its extension.
    """
    _, extension = os.path.splitext(file_path)
    extension = extension.lower()
    
    if extension == '.xlsx':
        return _process_xlsx(file_path)
    elif extension == '.xls':
        return _process_xls(file_path)
    elif extension == '.csv':
        return _process_csv(file_path)
    else:
        print(f"Unsupported spreadsheet format: {extension}")
        return ""
